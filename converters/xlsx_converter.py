from pathlib import Path
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from .base_converter import BaseConverter


class XLSXConverter(BaseConverter):
    def convert(self, input_path: Path, output_path: Path):
        pdfmetrics.registerFont(TTFont("DejaVu", "DejaVuSans.ttf"))

        wb = load_workbook(str(input_path), data_only=True)
        ws = wb.active

        lines = []
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(v) if v is not None else "" for v in row]
            lines.append(" | ".join(row_vals))

        text = "\n".join(lines)

        c = canvas.Canvas(str(output_path), pagesize=letter)
        c.setFont("DejaVu", 9)

        y = 750
        for line in text.split("\n"):
            c.drawString(50, y, line[:200])
            y -= 12
            if y < 50:
                c.showPage()
                c.setFont("DejaVu", 9)
                y = 750

        c.save()
